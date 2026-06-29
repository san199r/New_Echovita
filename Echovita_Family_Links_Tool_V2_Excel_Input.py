import os
import re
import math
import time
import subprocess
import sys

from openpyxl import Workbook, load_workbook
import undetected_chromedriver as uc
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException

import re
import subprocess


def _parse_major(version_text: str):
    # Accepts: "Google Chrome 144.0.7559.110" or "Chrome 144.0...."
    m = re.search(r"(\d+)\.", version_text or "")
    return int(m.group(1)) if m else None


def get_chrome_major_version():
    """
    Robust Windows Chrome major version detection:
    1) try: where chrome
    2) try common install paths (Program Files + LocalAppData)
    3) try registry version (BLBeacon)
    Returns: int major version (e.g., 144) or None
    """

    # --- 1) Try PATH: `where chrome`
    try:
        where_out = subprocess.check_output(["where", "chrome"], text=True, stderr=subprocess.STDOUT)
        for line in where_out.splitlines():
            chrome_path = line.strip()
            if chrome_path and chrome_path.lower().endswith("chrome.exe") and os.path.exists(chrome_path):
                try:
                    ver_out = subprocess.check_output([chrome_path, "--version"], text=True).strip()
                    major = _parse_major(ver_out)
                    if major:
                        return major
                except Exception:
                    pass
    except Exception:
        pass

    # --- 2) Try common paths (system + per-user)
    local = os.environ.get("LOCALAPPDATA", "")
    candidates = [
        r"C:\Program Files\Google\Chrome\Application\chrome.exe",
        r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
        os.path.join(local, r"Google\Chrome\Application\chrome.exe"),
        os.path.join(local, r"Google\Chrome Beta\Application\chrome.exe"),
        os.path.join(local, r"Google\Chrome SxS\Application\chrome.exe"),
    ]
    for chrome_path in candidates:
        if chrome_path and os.path.exists(chrome_path):
            try:
                ver_out = subprocess.check_output([chrome_path, "--version"], text=True).strip()
                major = _parse_major(ver_out)
                if major:
                    return major
            except Exception:
                pass

    # --- 3) Try registry (version stored here on many installs)
    reg_keys = [
        r"HKCU\Software\Google\Chrome\BLBeacon",
        r"HKLM\Software\Google\Chrome\BLBeacon",
        r"HKLM\Software\WOW6432Node\Google\Chrome\BLBeacon",
    ]
    for key in reg_keys:
        try:
            reg_out = subprocess.check_output(
                ["reg", "query", key, "/v", "version"],
                text=True,
                stderr=subprocess.STDOUT
            )
            # Example line: "version    REG_SZ    144.0.7559.110"
            m = re.search(r"version\s+REG_SZ\s+([0-9.]+)", reg_out, re.IGNORECASE)
            if m:
                major = _parse_major(m.group(1))
                if major:
                    return major
        except Exception:
            pass

    return None


PER_PAGE = 24
OUTPUT_FILENAME = "Echovita_Family_Links_Tool_Output.xlsx"
CAPTCHA_TEXT = (
    "This website uses a security service to protect against malicious bots. "
    "This page is displayed while the website verifies you are not a bot."
)


# ---------------------- TKINTER HELPERS ---------------------- #

def select_input_file():
    # If passed via command line
    if len(sys.argv) > 1:
        return sys.argv[1]
    # Default path for GitHub Actions
    return os.path.join(os.getcwd(), "workspace", "input", "Input.xlsx")

def show_info(title, message):
    print(f"INFO [{title}]: {message}")

def show_error(title, message):
    print(f"ERROR [{title}]: {message}")

# ---------------------- EXCEL HELPERS ---------------------- #

def get_output_file_path(input_file_path):
    output_dir = os.path.join(os.getcwd(), "workspace", "output")
    os.makedirs(output_dir, exist_ok=True)
    return os.path.join(output_dir, OUTPUT_FILENAME)


def create_or_load_workbook(output_path):
    if os.path.exists(output_path):
        try:
            return load_workbook(output_path)
        except Exception:
            raise Exception(
                f"Output file exists but is not a valid Excel workbook:\n{output_path}\n"
                f"Delete it and run again."
            )

    wb = Workbook()
    ws = wb.active
    ws.title = "temp"
    wb.save(output_path)
    return wb


def ensure_sheet(wb, sheet_name, current_url, output_path):
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_found = True
    else:
        ws = wb.create_sheet(title=sheet_name)
        sheet_found = False

        headers = [
            "S.no.", "County Name", "State", "Input URL", "Name", "DOD", "Age", "Record URL",
            "TotalCount", "LastSerial", "LastPage", "LastRecordIndex",
            "Total Pages", "Sheet URL",
            "Husband Name", "Wife Name",
            "Child 1 Name", "Child 2 Name", "Child 3 Name", "Child 4 Name", "Child 5 Name",
            "Parent 1 Name", "Parent 2 Name",
            "Sibling 1 Name", "Sibling 2 Name", "Sibling 3 Name", "Sibling 4 Name", "Sibling 5 Name",
            "Sister in Law 1 Name", "Sister in Law 2 Name",
            "Brother in Law 1 Name", "Brother in Law 2 Name",
            "Grandchild 1 Name", "Grandchild 2 Name", "Grandchild 3 Name", "Grandchild 4 Name", "Grandchild 5 Name",
            "Niece 1 Name", "Niece 2 Name",
            "Nephew 1 Name", "Nephew 2 Name"
        ]
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        ws.cell(row=2, column=14, value=current_url)

        if "temp" in wb.sheetnames and len(wb.sheetnames) > 1:
            temp_ws = wb["temp"]
            wb.remove(temp_ws)

        wb.save(output_path)

    return ws, sheet_found


def read_int_cell(ws, row, col, default=0):
    value = ws.cell(row=row, column=col).value
    if value is None or value == "":
        return default
    if str(value).strip() == "Page not found":
        return default
    try:
        return int(float(str(value).strip()))
    except Exception:
        return default


def read_text_cell(ws, row, col, default=""):
    value = ws.cell(row=row, column=col).value
    if value is None:
        return default
    return str(value)


def save_workbook_safe(wb, output_path):
    wb.save(output_path)


def read_input_excel(input_file_path):
    """
    Reads input Excel.
    Expected columns:
    Col 1 = S NO
    Col 2 = COUNTY NAME
    Col 3 = STATE
    Col 4 = ECHOVITA LINKS

    Header row is ignored.
    """
    wb_in = load_workbook(input_file_path, data_only=True)
    ws_in = wb_in.active

    input_rows = []

    for row in range(2, ws_in.max_row + 1):
        sno = ws_in.cell(row=row, column=1).value
        county_name = ws_in.cell(row=row, column=2).value
        state = ws_in.cell(row=row, column=3).value
        echovita_link = ws_in.cell(row=row, column=4).value

        if echovita_link is None or str(echovita_link).strip() == "":
            continue

        input_rows.append({
            "sno": sno,
            "county_name": "" if county_name is None else str(county_name).strip(),
            "state": "" if state is None else str(state).strip(),
            "url": str(echovita_link).strip()
        })

    return input_rows


# ---------------------- SELENIUM / PAGE HELPERS ---------------------- #

import os


def create_driver():
    chrome_major = get_chrome_major_version()
    if not chrome_major:
        print("Could not detect your Google Chrome version.")
        raise Exception("Could not detect Google Chrome version.")

    options = uc.ChromeOptions()
    options.add_argument("--start-maximized")
    options.add_argument("--log-level=3")
    options.add_argument("--disable-logging")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--disable-background-timer-throttling")
    options.add_argument("--disable-backgrounding-occluded-windows")
    options.add_argument("--disable-renderer-backgrounding")
    options.add_argument("--disable-features=CalculateNativeWinOcclusion")
    
    # Required for GitHub Actions headless execution to prevent hanging
    options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-gpu")
    options.add_argument("--disable-dev-shm-usage")

    driver = uc.Chrome(options=options, version_main=chrome_major)
    driver.set_page_load_timeout(60)
    return driver


def get_body_text(driver):
    try:
        return driver.find_element("tag name", "body").text or ""
    except Exception:
        return ""


def handle_captcha_if_present(driver):
    body_text = get_body_text(driver)
    if CAPTCHA_TEXT in body_text:
        print("ERROR: Captcha Detected! Cannot solve in headless mode.")
        # Raise exception to skip this page
        raise Exception("CAPTCHA_DETECTED")


def open_url(driver, url, expect_text=None, timeout=60):
    driver.get(url)
    time.sleep(3)
    handle_captcha_if_present(driver)

    if expect_text:
        def condition(d):
            body = get_body_text(d).lower()

            server_error_words = [
                "oops! an error occurred",
                "http error",
                "this page isn’t working",
                "this page isn't working",
                "this site can’t be reached",
                "this site can't be reached",
                "server error",
                "service unavailable",
                "bad gateway",
                "gateway timeout",
                "internal server error",
                "err_http",
                "err_connection",
                "err_timed_out"
            ]

            return (
                expect_text.lower() in body
                or "page not found" in body
                or any(word in body for word in server_error_words)
            )

        WebDriverWait(driver, timeout).until(condition)
        handle_captcha_if_present(driver)


def page_contains_text(driver, text):
    return text.lower() in get_body_text(driver).lower()


def handle_server_or_page_error(driver, ws, wb, output_path):
    body_text = get_body_text(driver)

    # Case 1: Echovita page error - write Error and move on
    if "oops! an error occurred" in body_text.lower():
        print("Oops error found. Writing Error and moving to next URL.")
        ws.cell(row=2, column=9, value="Error")
        save_workbook_safe(wb, output_path)
        return "ERROR_MOVE_NEXT"

    # Case 2: HTTPS / server error
    server_error_words = [
        "HTTP ERROR",
        "This page isn’t working",
        "This site can’t be reached",
        "server error",
        "Service Unavailable",
        "Bad Gateway",
        "Gateway Timeout",
        "Internal Server Error",
        "ERR_HTTP",
        "ERR_CONNECTION",
        "ERR_TIMED_OUT"
    ]

    if any(word.lower() in body_text.lower() for word in server_error_words):
        print("ERROR: HTTPS/server error found. Cannot manually reload in headless mode.")
        return "ERROR_MOVE_NEXT"

    return "OK"


def build_page_url(current_url, current_page):
    if re.search(r'([?&])page=\d+', current_url, flags=re.I):
        return re.sub(r'([?&])page=\d+', rf'\1page={current_page}', current_url, flags=re.I)

    if "?" in current_url:
        return f"{current_url}&page={current_page}"
    return f"{current_url}?page={current_page}"


# ---------------------- JAVASCRIPT SCRAPERS ---------------------- #

JS_TOTAL_COUNT = r"""
function scrapeObitCount() {
  const root = document.querySelector('#desktopObitSearch');
  if (!root) return '';
  const el = root.querySelector('div.d-flex.flex-column > span.eh-font-weight400')
           || root.querySelector('div.d-flex.flex-column > span');
  const txt = (el?.textContent || '').replace(/\s+/g, ' ').trim();
  return txt;
}
return scrapeObitCount();
"""

JS_RECORD_LINKS = r"""
function getRecordLinks() {
  var origin = location.origin;
  var nodes = document.querySelectorAll('.obit-list-wrapper a.text-name-obit-in-list[href]');
  if (!nodes || !nodes.length) return '';

  var seen = Object.create(null), out = [];
  for (var i = 0; i < nodes.length; i++) {
    var href = (nodes[i].getAttribute('href') || '').trim();
    if (!href) continue;
    if (href.charAt(0) === '/') href = origin + href;
    if (!seen[href]) { seen[href] = true; out.push(href); }
  }

  return out.join('|') + '';
}
return getRecordLinks();
"""

JS_NAME = r"""
function getName() {
  try {
    var el =
      document.querySelector('.obit-main-info-wrapper-min-height p.my-auto.h1.text-white.font-weight-bolder')
      || document.querySelector('h1.text-center.text-lg-left.font-weight-bolder.mb-2')
      || document.querySelector('h1, .display-4.font-weight-bolder');

    var txt = (el && el.textContent || '').replace(/\s+/g, ' ').trim();
    return txt + '';
  } catch (e) {
    return '';
  }
}
return getName();
"""

JS_DOD = r"""
function getDOD() {
  try {
    var p = document.querySelector('.obit-main-info-wrapper-min-height p.mt-2.mb-1.text-white.font-weight-bold');
    var raw = (p && p.textContent || '').replace(/\s+/g, ' ').trim();
    if (!raw) return '';

    var dates = raw.match(/([A-Za-z]+ \d{1,2}, \d{4})/g);
    if (dates && dates.length) {
      return (dates[dates.length - 1] + '');
    }
    return '';
  } catch (e) {
    return '';
  }
}
return getDOD();
"""

JS_AGE = r"""
function getAge() {
  try {
    var p = document.querySelector('.obit-main-info-wrapper-min-height p.mt-2.mb-1.text-white.font-weight-bold');
    var raw = (p && p.textContent || '').replace(/\s+/g, ' ').trim();
    if (!raw) return '';

    var m = raw.match(/\((\d{1,3})\s*years?\s*old\)/i);
    if (m) return (m[1] + '');

    m = raw.match(/\b(\d{1,3})\s*years?\b/i);
    if (m) return (m[1] + '');

    return '';
  } catch (e) {
    return '';
  }
}
return getAge();
"""

JS_RELATIONS = r"""
function extractRelations() {
  function getObituaryText() {
    var el = document.querySelector('#obituary');
    if (!el) return '';
    var txt = el.innerText || el.textContent || '';
    txt = txt.replace(/\s+/g, ' ').trim();
    return txt;
  }

  var text = getObituaryText();
  if (!text) {
    return new Array(27).fill('').join('|');
  }

  var SENT_BOUNDARY =
    '(?:;|\\.(?=\\s+(?:He|She|Visitation|Calling|The|A\\b|An\\b|Born|Funeral|Service\\b|Services\\b|Mass|Interment|Burial|Cremation|Friends?|Family|Memorial|Obituary|Graveside|There\\b|Arrangements?|A\\s+celebration|Expressions?|Donations?|In\\b|Online|Condolences?))|$)';

  function cleanName(s) {
    if (!s) return '';
    s = s.replace(/\s+of\s+[^,;.]+(?:,[^,;.]+)*/gi, '');
    s = s.replace(/^[-\s,;:.]+/, '');
    s = s.replace(/[-\s,;:.]+$/, '');
    return s.trim();
  }

  function nameLooksLikeName(s) {
    if (!s) return false;
    if (!/[A-Za-z]/.test(s)) return false;
    if (/\d/.test(s)) return false;
    if (/\b(grandchildren?|grandchild|granddaughters?|grandsons?|sisters?|brothers?|children?|sons?|daughters?|nieces?|nephews?)\b/i.test(s)) {
      return false;
    }
    if (/\b(and\/or|Humane\s+Society|Hospice|Hospital|Funeral\s+Home|Funeral\b|Cemetery|Church|Foundation|Memorial|Lexington|donations?|contributions?|expressions?\s+of\s+sympathy|online\s+condolences?)\b/i.test(s)) {
      return false;
    }
    if (s.indexOf(' ') === -1 && s.indexOf('-') === -1 && s.length <= 2) {
      return false;
    }
    return true;
  }

  function splitNameList(str) {
    if (!str) return [];
    var t = str.replace(/\band\b/gi, ',');
    var parts = t.split(',');
    var names = [];
    for (var i = 0; i < parts.length; i++) {
      var p = parts[i].trim();
      if (!p) continue;

      p = p.replace(/\s+of\s+[^,;.]+(?:,[^,;.]+)*/gi, '');

      p = p.replace(
        /^(his|her)\s+(son|sons|daughter|daughters|child|children|grandchild|grandchildren|granddaughter|granddaughters|grandson|grandsons|niece|nieces|nephew|nephews|brother|brothers|sister|sisters)\s+/i,
        ''
      );

      var cleaned = cleanName(p);
      if (!cleaned) continue;
      if (!nameLooksLikeName(cleaned)) continue;

      names.push(cleaned);
    }
    return names;
  }

  var husbandName = '';
  var wifeName = '';

  var mH = text.match(new RegExp(
    '\\bher\\s+(?:husband|spouse)\\s*,?\\s+(.+?)(?=' + SENT_BOUNDARY + ')',
    'i'
  ));
  if (mH) husbandName = cleanName(mH[1]);

  var mW = text.match(new RegExp(
    '\\bhis\\s+(?:wife|spouse)\\s*,?\\s+(.+?)(?=' + SENT_BOUNDARY + ')',
    'i'
  ));
  if (mW) wifeName = cleanName(mW[1]);

  var parent1 = '';
  var parent2 = '';

  var mP = text.match(new RegExp(
    '(?:his|her)\\s+parents,\\s*(.+?)\\s+(?:and|&)\\s+(.+?)(?=' + SENT_BOUNDARY + ')',
    'i'
  ));
  if (mP) {
    parent1 = cleanName(mP[1]);
    parent2 = cleanName(mP[2]);
  }

  if (!parent1) {
    var fMatch = text.match(new RegExp(
      '(?:his|her)\\s+father\\s+(.+?)(?=' + SENT_BOUNDARY + ')',
      'i'
    ));
    if (fMatch) parent1 = cleanName(fMatch[1]);
  }

  if (!parent2) {
    var mMatch2 = text.match(new RegExp(
      '(?:his|her)\\s+mother\\s+(.+?)(?=' + SENT_BOUNDARY + ')',
      'i'
    ));
    if (mMatch2) parent2 = cleanName(mMatch2[1]);
  }

  var sisters = [];
  var brothers = [];
  var siblings = [];

  var mSibBlock = text.match(new RegExp(
    '(?:his|her)\\s+siblings?,\\s*(.+?)(?=' + SENT_BOUNDARY + ')',
    'i'
  ));
  if (mSibBlock) siblings = siblings.concat(splitNameList(mSibBlock[1]));

  var mSisBlock = text.match(new RegExp(
    '(?:his|her)\\s+sisters?(?!-in-law)\\s*,\\s*(.+?)(?=' + SENT_BOUNDARY + ')',
    'i'
  ));
  if (mSisBlock) sisters = sisters.concat(splitNameList(mSisBlock[1]));

  var ms;
  var reSingSis = new RegExp(
    '(?:his|her)\\s+sister(?!-in-law)\\s+(.+?)(?=' + SENT_BOUNDARY + ')',
    'ig'
  );
  while ((ms = reSingSis.exec(text)) && sisters.length < 5) {
    var nm = cleanName(ms[1]);
    if (nameLooksLikeName(nm)) sisters.push(nm);
  }

  var mBroBlock = text.match(new RegExp(
    '(?:his|her)\\s+brothers?(?!-in-law)\\s*,\\s*(.+?)(?=' + SENT_BOUNDARY + ')',
    'i'
  ));
  if (mBroBlock) brothers = brothers.concat(splitNameList(mBroBlock[1]));

  var reSingBro = new RegExp(
    '(?:his|her)\\s+brother(?!-in-law)\\s+(.+?)(?=' + SENT_BOUNDARY + ')',
    'ig'
  );
  while ((ms = reSingBro.exec(text)) && brothers.length < 5) {
    var nb = cleanName(ms[1]);
    if (nameLooksLikeName(nb)) brothers.push(nb);
  }

  siblings = siblings.concat(sisters, brothers);

  var seenSib = Object.create(null);
  var sibFinal = [];
  for (var i = 0; i < siblings.length && sibFinal.length < 5; i++) {
    var s = siblings[i];
    if (!s) continue;
    if (seenSib[s]) continue;
    seenSib[s] = true;
    sibFinal.push(s);
  }

  var children = [];

  var reChildBlock = new RegExp(
    '(?:his|her)\\s+(?:children?|sons?|daughters?)\\s*,\\s*(.+?)(?=' + SENT_BOUNDARY + ')',
    'ig'
  );
  while ((ms = reChildBlock.exec(text))) {
    children = children.concat(splitNameList(ms[1]));
  }

  var reChild = new RegExp(
    '(?:his|her)\\s+(?:son|daughter)\\s+(.+?)(?=' + SENT_BOUNDARY + ')',
    'ig'
  );
  while ((ms = reChild.exec(text))) {
    var c = cleanName(ms[1]);
    if (nameLooksLikeName(c)) children.push(c);
  }

  var seenChild = Object.create(null);
  var childFinal = [];
  for (var j = 0; j < children.length && childFinal.length < 5; j++) {
    var ch = children[j];
    if (!ch) continue;
    if (seenChild[ch]) continue;
    seenChild[ch] = true;
    childFinal.push(ch);
  }

  var grandkids = [];

  var reGrandBlock = new RegExp(
    '(?:his|her)\\s+grand(?:children?|child|daughters?|sons?)\\s*,\\s*(.+?)(?=' + SENT_BOUNDARY + ')',
    'ig'
  );
  while ((ms = reGrandBlock.exec(text))) {
    grandkids = grandkids.concat(splitNameList(ms[1]));
  }

  var reGrandSingle = new RegExp(
    '(?:his|her)\\s+grand(?:son|daughter)\\s+(.+?)(?=' + SENT_BOUNDARY + ')',
    'ig'
  );
  while ((ms = reGrandSingle.exec(text))) {
    var g = cleanName(ms[1]);
    if (nameLooksLikeName(g)) grandkids.push(g);
  }

  if (grandkids.length > 5) grandkids = grandkids.slice(0, 5);

  var nieces = [];
  var nephews = [];

  var reNieceBlock = new RegExp(
    '(?:his|her)\\s+nieces?,\\s*(.+?)(?=' + SENT_BOUNDARY + ')',
    'ig'
  );
  while ((ms = reNieceBlock.exec(text))) {
    nieces = nieces.concat(splitNameList(ms[1]));
  }

  var reNiece = new RegExp(
    '(?:his|her)\\s+niece\\s+(.+?)(?=' + SENT_BOUNDARY + ')',
    'ig'
  );
  while ((ms = reNiece.exec(text))) {
    var n = cleanName(ms[1]);
    if (nameLooksLikeName(n)) nieces.push(n);
  }

  var reNephewBlock = new RegExp(
    '(?:his|her)\\s+nephews?,\\s*(.+?)(?=' + SENT_BOUNDARY + ')',
    'ig'
  );
  while ((ms = reNephewBlock.exec(text))) {
    nephews = nephews.concat(splitNameList(ms[1]));
  }

  var reNephew = new RegExp(
    '(?:his|her)\\s+nephew\\s+(.+?)(?=' + SENT_BOUNDARY + ')',
    'ig'
  );
  while ((ms = reNephew.exec(text))) {
    var np = cleanName(ms[1]);
    if (nameLooksLikeName(np)) nephews.push(np);
  }

  if (nieces.length > 2) nieces = nieces.slice(0, 2);
  if (nephews.length > 2) nephews = nephews.slice(0, 2);

  var sisterInLaws = [];
  var brotherInLaws = [];

  var reSilBlock = new RegExp(
    'sisters?-in-law\\s*,\\s*(.+?)(?=' + SENT_BOUNDARY + ')',
    'ig'
  );
  while ((ms = reSilBlock.exec(text))) {
    sisterInLaws = sisterInLaws.concat(splitNameList(ms[1]));
  }

  var reSil = new RegExp(
    'sister-in-law\\s+(.+?)(?=' + SENT_BOUNDARY + ')',
    'ig'
  );
  while ((ms = reSil.exec(text))) {
    var sil = cleanName(ms[1]);
    if (nameLooksLikeName(sil)) sisterInLaws.push(sil);
  }

  var reBilBlock = new RegExp(
    'brothers?-in-law\\s*,\\s*(.+?)(?=' + SENT_BOUNDARY + ')',
    'ig'
  );
  while ((ms = reBilBlock.exec(text))) {
    brotherInLaws = brotherInLaws.concat(splitNameList(ms[1]));
  }

  var reBil = new RegExp(
    'brother-in-law\\s+(.+?)(?=' + SENT_BOUNDARY + ')',
    'ig'
  );
  while ((ms = reBil.exec(text))) {
    var bil = cleanName(ms[1]);
    if (nameLooksLikeName(bil)) brotherInLaws.push(bil);
  }

  if (sisterInLaws.length > 2) sisterInLaws = sisterInLaws.slice(0, 2);
  if (brotherInLaws.length > 2) brotherInLaws = brotherInLaws.slice(0, 2);

  var fields = [
    husbandName,
    wifeName,
    childFinal[0] || '',
    childFinal[1] || '',
    childFinal[2] || '',
    childFinal[3] || '',
    childFinal[4] || '',
    parent1,
    parent2,
    sibFinal[0] || '',
    sibFinal[1] || '',
    sibFinal[2] || '',
    sibFinal[3] || '',
    sibFinal[4] || '',
    sisterInLaws[0] || '',
    sisterInLaws[1] || '',
    brotherInLaws[0] || '',
    brotherInLaws[1] || '',
    grandkids[0] || '',
    grandkids[1] || '',
    grandkids[2] || '',
    grandkids[3] || '',
    grandkids[4] || '',
    nieces[0] || '',
    nieces[1] || '',
    nephews[0] || '',
    nephews[1] || ''
  ];

  return fields.join('|');
}
return extractRelations();
"""


def get_total_count(driver):
    label = driver.execute_script(JS_TOTAL_COUNT) or ""
    match = re.search(r'(\d[\d,]*)', str(label))
    if not match:
        return 0
    try:
        return int(match.group(1).replace(",", ""))
    except Exception:
        return 0


def get_record_links(driver):
    raw = driver.execute_script(JS_RECORD_LINKS) or ""
    raw = str(raw).strip()
    if not raw:
        return []
    return [x.strip() for x in raw.split("|") if x.strip()]


def get_name(driver):
    return (driver.execute_script(JS_NAME) or "").strip()


def get_dod(driver):
    return (driver.execute_script(JS_DOD) or "").strip()


def get_age(driver):
    return (driver.execute_script(JS_AGE) or "").strip()


def get_relations(driver):
    raw = driver.execute_script(JS_RELATIONS) or ""
    parts = [x.strip() for x in str(raw).split("|")]
    while len(parts) < 27:
        parts.append("")
    return parts[:27]


# ---------------------- MAIN LOGIC ---------------------- #

def run():
    driver = None
    wb = None
    output_path = None

    try:
        input_file_path = select_input_file()
        if not input_file_path:
            show_info("Cancelled", "No input file selected.")
            return

        output_path = get_output_file_path(input_file_path)
        wb = create_or_load_workbook(output_path)

        input_rows = read_input_excel(input_file_path)

        if not input_rows:
            show_info("No URLs", "No Echovita links found in column 4.")
            return

        print(f"Total input URLs found: {len(input_rows)}")

        serial = 1

        for url_index, input_data in enumerate(input_rows, start=1):
            current_url = input_data["url"]
            county_name = input_data["county_name"]
            state = input_data["state"]

            print(f"\nProcessing Url {url_index}: {current_url}")

            sheet_name = str(url_index)
            ws, sheet_found = ensure_sheet(wb, sheet_name, current_url, output_path)

            if not sheet_found:
                next_write_row = 2
                last_serial_number = 0
            else:
                last_serial_stored_cell = ws.cell(row=2, column=10).value
                if last_serial_stored_cell in (None, ""):
                    last_serial_number = 0
                else:
                    try:
                        last_serial_number = int(float(str(last_serial_stored_cell).strip()))
                    except Exception:
                        last_serial_number = 0

                next_write_row = last_serial_number + 2
                if next_write_row < 2:
                    next_write_row = 2

            sheet_url_cell = ws.cell(row=2, column=14).value
            if sheet_url_cell in (None, ""):
                ws.cell(row=2, column=14, value=current_url)
                save_workbook_safe(wb, output_path)
                sheet_url = current_url
            else:
                sheet_url = str(sheet_url_cell)

            total_count_for_url = read_int_cell(ws, 2, 9, 0)
            last_serial_stored = read_int_cell(ws, 2, 10, 0)
            last_page_stored = read_int_cell(ws, 2, 11, 0)
            last_record_index_stored = read_int_cell(ws, 2, 12, 0)

            total_count_raw = read_text_cell(ws, 2, 9, "")
            if total_count_raw == "Page not found":
                total_count_for_url = 0

            if last_serial_number < last_serial_stored:
                last_serial_number = last_serial_stored

            if driver is None:
                driver = create_driver()

            skip_this_url = False
            try:
                open_url(driver, current_url, expect_text="Receive obituaries", timeout=120)
                while True:
                    error_status = handle_server_or_page_error(driver, ws, wb, output_path)

                    if error_status == "ERROR_MOVE_NEXT":
                        skip_this_url = True
                        break

                    if error_status == "RECHECK":
                        handle_captcha_if_present(driver)
                        continue

                    break
            except TimeoutException:

                if page_contains_text(driver, "Page not found"):
                    print(f"Url {url_index}: Page not found")
                    ws.cell(row=2, column=9, value="Page not found")
                    save_workbook_safe(wb, output_path)
                    continue
                raise

            if skip_this_url:
                continue

            if page_contains_text(driver, "Page not found"):
                print(f"Url {url_index}: Page not found")
                ws.cell(row=2, column=9, value="Page not found")
                save_workbook_safe(wb, output_path)
                continue

            if total_count_for_url == 0:
                total_count_for_url = get_total_count(driver)
                ws.cell(row=2, column=9, value=total_count_for_url)
                save_workbook_safe(wb, output_path)
                print(f"Found {total_count_for_url} records in Url {url_index}")
                time.sleep(2)
            else:
                print(f"Found {total_count_for_url} records in Url {url_index}")

            skip_this_url = False
            if total_count_for_url > 0 and last_serial_number >= total_count_for_url:
                ws.cell(row=2, column=10, value=last_serial_number)
                save_workbook_safe(wb, output_path)
                print(f"Url {url_index} already completed. Skipping.")
                skip_this_url = True

            if skip_this_url:
                continue

            start_page = last_page_stored if last_page_stored > 0 else 1
            start_record_index = last_record_index_stored + 1 if last_record_index_stored > 0 else 1

            current_page = start_page
            record_serial = last_serial_number + 1

            total_pages = 0 if total_count_for_url <= 0 else math.ceil(total_count_for_url / PER_PAGE)
            ws.cell(row=2, column=13, value=total_pages)
            save_workbook_safe(wb, output_path)

            if total_pages <= 0:
                continue

            for current_page in range(1, total_pages + 1):
                if current_page < start_page:
                    continue

                page_url = build_page_url(current_url, current_page)
                open_url(driver, page_url, expect_text="Receive obituaries", timeout=120)
                while True:
                    error_status = handle_server_or_page_error(driver, ws, wb, output_path)

                    if error_status == "ERROR_MOVE_NEXT":
                        skip_this_url = True
                        break

                    if error_status == "RECHECK":
                        handle_captcha_if_present(driver)
                        continue

                    break

                if skip_this_url:
                    break

                print(f"Opened page {current_page} for Url {url_index}")

                records_list = get_record_links(driver)
                print(f"Found {len(records_list)} record links on page {current_page} for Url {url_index}")
                record_index = 1

                for record_url in records_list:
                    if current_page == start_page and record_index < start_record_index:
                        record_index += 1
                        continue

                    open_url(driver, record_url, expect_text="Obituary", timeout=120)

                    name_text = get_name(driver)
                    dod_text = get_dod(driver)
                    age_text = get_age(driver)
                    rel_fields = get_relations(driver)

                    ws.cell(row=next_write_row, column=1, value=record_serial)
                    ws.cell(row=next_write_row, column=2, value=county_name)
                    ws.cell(row=next_write_row, column=3, value=state)
                    ws.cell(row=next_write_row, column=4, value=sheet_url)
                    ws.cell(row=next_write_row, column=5, value=name_text)
                    ws.cell(row=next_write_row, column=6, value=dod_text)
                    ws.cell(row=next_write_row, column=7, value=age_text)
                    ws.cell(row=next_write_row, column=8, value=record_url)

                    for idx, value in enumerate(rel_fields, start=15):
                        ws.cell(row=next_write_row, column=idx, value=value)

                    ws.cell(row=2, column=10, value=record_serial)
                    ws.cell(row=2, column=11, value=current_page)
                    ws.cell(row=2, column=12, value=record_index)

                    save_workbook_safe(wb, output_path)
                    print(f"Saved Record {record_serial} with name {name_text}")

                    record_serial += 1
                    next_write_row += 1
                    record_index += 1

            serial += 1

        save_workbook_safe(wb, output_path)
        show_info("All URLs Done", f"Output saved in excel:\n{output_path}")

    except Exception as e:
        err_text = (
            f"An error occurred.\n\n"
            f"Error: {str(e)}\n\n"
            f"The tool will now close safely."
        )
        print(f"ERROR: {str(e)}")
        show_error("Tool Error", err_text)

    finally:
        try:
            if wb and output_path:
                wb.save(output_path)
        except Exception:
            pass

        try:
            if driver:
                driver.quit()
        except Exception:
            pass


if __name__ == "__main__":
    run()